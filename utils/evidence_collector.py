import json
from collections import defaultdict
from typing import Any, Dict, List

class EvidenceCollector:
    def __init__(self):
        # Data structure: { "TC-1": {"api": [...], "db": [...]}, ... }
        self.evidences = defaultdict(lambda: {"api": [], "db": [], "tc_name": "", "expected_result": "", "precondition": "", "status": "Passed"})

    def update_excel_status(self, tc_id: str, status: str, test_data: str, expected_result: str = ""):
        try:
            import openpyxl
            from datetime import datetime
            wb = openpyxl.load_workbook('collections/test_script.xlsx')
            ws = wb.active
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=4).value == tc_id: # Col 4 is TC-ID
                    ws.cell(row=row, column=11).value = status # Col 11 is Status
                    if test_data:
                        ws.cell(row=row, column=9).value = test_data # Col 9 is Test Data
                    if expected_result:
                        ws.cell(row=row, column=10).value = expected_result # Col 10 is Expected Result
                    ws.cell(row=row, column=14).value = datetime.now().strftime("%d/%m/%Y") # Col 14 is Tanggal Testing
                    break
            wb.save('collections/test_script.xlsx')
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Failed to update excel: {e}")

    def set_test_metadata(self, tc_id: str, tc_name: str, expected_result: str, precondition: str = ""):
        self.evidences[tc_id]["tc_name"] = tc_name
        self.evidences[tc_id]["expected_result"] = expected_result
        self.evidences[tc_id]["precondition"] = precondition

    def set_test_status(self, tc_id: str, status: str):
        self.evidences[tc_id]["status"] = status
        
        # Use custom test data if provided by the test case
        if self.evidences[tc_id].get("custom_test_data"):
            test_data_str = self.evidences[tc_id]["custom_test_data"]
        else:
            # Build test data string from all payloads sent in this test case
            test_data_str = ""
            for api in self.evidences[tc_id].get("api", []):
                payload = api.get("request_payload", {})
                if payload:
                    # Add important fields as a summary
                    if "nomor_transaksi" in payload and f"nomor_transaksi: {payload['nomor_transaksi']}" not in test_data_str:
                        test_data_str += f"nomor_transaksi: {payload['nomor_transaksi']}\n"
                    if "nomor_loan" in payload and f"nomor_loan: {payload['nomor_loan']}" not in test_data_str:
                        test_data_str += f"nomor_loan: {payload['nomor_loan']}\n"
                    if "ktp" in payload and f"ktp: {payload['ktp']}" not in test_data_str:
                        test_data_str += f"ktp: {payload['ktp']}\n"
        
        # fallback to full json if summary is empty but there's a payload
        if not test_data_str and self.evidences[tc_id].get("api"):
            payload = self.evidences[tc_id]["api"][0].get("request_payload", {})
            if payload:
                import json
                test_data_str = json.dumps(payload, indent=2)
                
        expected_result = self.evidences[tc_id].get("expected_result", "")
        self.update_excel_status(tc_id, status, test_data_str, expected_result)

    def add_api_evidence(self, tc_id: str, url: str, method: str, request_payload: dict, response_json: dict, status_code: int):
        self.evidences[tc_id]["api"].append({
            "url": url,
            "method": method,
            "request_payload": request_payload,
            "response_json": response_json,
            "status_code": status_code
        })

    def add_db_evidence(self, tc_id: str, query: str, result: Any):
        self.evidences[tc_id]["db"].append({
            "query": query,
            "result": result
        })

    def add_epolis_evidence(self, tc_id: str, qr_result: str, image_paths: List[str]):
        self.evidences[tc_id]["epolis"] = {
            "qr_result": qr_result,
            "image_paths": image_paths
        }

    def add_ui_evidence(self, tc_id: str, system_name: str, screenshot_path: str):
        if "ui" not in self.evidences[tc_id]:
            self.evidences[tc_id]["ui"] = []
        self.evidences[tc_id]["ui"].append({
            "system_name": system_name,
            "screenshot_path": screenshot_path
        })

    def get_all_evidences(self) -> dict:
        return dict(self.evidences)

    def clear(self):
        self.evidences.clear()

evidence_collector = EvidenceCollector()
